import datetime
from openpyxl import load_workbook, Workbook
import os
import re


def is_watch_group(contact, group_name):
    if contact.nick == group_name:
        return True
    else:
        return False


def dialog_clearify(content):
    """
    这里面的顺序不要乱
    :param content:
    :return:
    """
    d = content
    if not isinstance(d, type("字符串类型")):
        return ""
    d = d.replace("  ", " ")
    # d = d.replace(" ", "")
    d = d.replace("~~~", "-")
    d = d.replace("~~", "-")
    d = d.replace("~", "-")
    d = d.replace("---", "-")
    d = d.replace("--", "-")
    d = d.replace("——", "-")
    d = d.replace("：", ":")
    d = d.replace("12层", "12楼")
    d = d.replace("13层", "13楼")
    d = d.replace("今日", "今天")
    d = d.replace("明日", "明天")
    d = d.replace("合昌", "和昌")
    d = d.replace("合唱", "和昌")
    d = d.replace("和唱", "和昌")
    d = d.replace("点半", ":30")
    d = d.replace("点30", ":30")
    d = d.replace("点三十", ":30")
    d = d.replace("点", ":00")
    d = d.replace("十一", "11")
    d = d.replace("十二", "12")
    d = d.replace("十三", "13")
    d = d.replace("十四", "14")
    d = d.replace("十五", "15")
    d = d.replace("十六", "16")
    d = d.replace("十七", "17")
    d = d.replace("十八", "18")
    d = d.replace("十九", "19")
    d = d.replace("二十", "20")
    d = d.replace("十", "10")
    d = d.replace("九", "9")
    d = d.replace("八", "8")
    d = d.replace("七", "7")
    d = d.replace("六", "6")
    d = d.replace("五", "5")
    d = d.replace("四", "4")
    d = d.replace("三", "3")
    d = d.replace("二", "2")
    d = d.replace("一", "1")
    d = d.replace("下午1:", "13:")
    d = d.replace("下午2:", "14:")
    d = d.replace("下午3:", "15:")
    d = d.replace("下午4:", "16:")
    d = d.replace("下午5:", "17:")
    d = d.replace("下午6:", "18:")
    d = d.replace("下午7:", "19:")
    d = d.replace("下午8:", "20:")
    d = d.replace("预订", "预定")
    d = d.replace("到", "-")
    d = d.replace("（", "(")
    d = d.replace("）", ")")
    d = d.replace("，", ".")
    d = d.replace(",", ".")
    d = d.replace("。", ".")
    return d


def is_cmd(dialog):
    if not isinstance(dialog, type("")):
        return ""
    if dialog.find("预定") > -1 and dialog.find("会议室") > -1:
        return dialog


def read_dialog(dialog):
    book_ornot = -1
    data_ans = datetime.date.today()
    meeting_room = ""
    start_time = ""
    end_time = ""
    keyword = {"预定或取消": ["预定", "取消"],
               "今天或明天": ["今天", "明天", "后天", "号", "日"],  # todo 处理 号 日
               "上下午": ["上午", "下午"],
               "开始时间": [":"],
               "结束时间": [":"],
               "是否和昌": ["和昌"],
               "楼层": ["12楼", "13楼"],
               "会议室": ["小会议室", "大会议室", "会议室"]}
    start_time_pos = 0
    for key in keyword:
        yet = False
        for value in keyword[key]:
            # print(value)
            if yet:
                break
            if -1 < dialog.find(value):
                if "今天或明天" == key:
                    if "明" == value[0]:
                        data_ans = data_ans+datetime.timedelta(days=1)
                    elif "后" == value[0]:
                        data_ans = data_ans + datetime.timedelta(days=2)
                    elif "号" == value[0] or "日" == value[0]:
                        pass
                    yet = True
                    continue
                elif "开始时间" == key:
                    start_time_pos = dialog.find(value)  # 冒号位置
                    start_time = dialog[start_time_pos - 2: start_time_pos + 3]
                    start_time = wash(start_time)
                    yet = True
                    continue
                elif "结束时间" == key:
                    end_time_pos = dialog.find(value, start_time_pos + 1)
                    end_time = dialog[end_time_pos - 2: end_time_pos + 3]
                    end_time = wash(end_time)
                    yet = True
                    continue
                elif "楼层" == key:
                    meeting_room = value
                    yet = True
                    continue
                elif "会议室" == key:
                    meeting_room = meeting_room + value
                    yet = True
                    continue
                elif "预定或取消" == key:
                    book_ornot = (value == "预定").__int__()  # 1 预定
                    yet = True
                    continue
                else:
                    yet = True
                    continue

    print("中间信息打印：", "book_ornot=", book_ornot)
    print("中间信息打印：", "data_ans.__str__()=", data_ans.__str__())
    print("中间信息打印：", "meeting_room=", meeting_room)
    print("中间信息打印：", "start_time=", start_time)
    print("中间信息打印：", "end_time=", end_time)
    return book_ornot, data_ans.__str__(), meeting_room, start_time, end_time


def find_fangjian(dialog):
    meeting_room = ""
    keyword = {"楼层": ["12楼", "13楼"],
               "会议室": ["小会议室", "大会议室", "会议室"]}
    for key in keyword:
        yet = False
        for value in keyword[key]:
            if yet:
                break
            if -1 < dialog.find(value):
                if "楼层" == key:
                    meeting_room += value
                    yet = True
                    continue
                elif "会议室" == key:
                    meeting_room += value
                    yet = True
                    continue
    return meeting_room


def find_shijian(dialog):
    st, et = None, None
    st_pos = dialog.find(":")  # 冒号位置
    if st_pos != -1:
        st = dialog[st_pos - 2: st_pos + 3]
        st = wash(st)
    et_pos = dialog.rfind(":")
    if st_pos != -1:
        et = dialog[et_pos - 2: et_pos + 3]
        et = wash(et)
    if isinstance(type(st), type("字符串")) and isinstance(type(et), type("字符串")):
        q = int(st[:st.find(":")])
        w = int(st[st.find(":") + 1:])
        e = int(et[:et.find(":")])
        r = int(et[et.find(":") + 1:])
        if e < q:
            e += 12
            et = str(e) + et[et.find(":"):]
        if 0 < w < 59 and 0 < r < 59:
            pass

    else:
        print("时间出错")
    return st, et


def find_yuding(dialog):
    if -1 < dialog.find("取消") < 5:
        return False
    else:
        if -1 < dialog.find("预定") < 5:
            return True


def find_riqi(dialog):
    findre = re.findall(r'([12]?[0-9]{1,2}[月.][123]?[0-9][日号])', dialog)  # 12月02日
    if len(findre) > 0:
        findre2 = findre[0]
        if findre2.find("月") == 1 or findre2.find(".") == 1:
            findre2 = "0" + findre2
        findre2 = findre2.replace("月", "-")
        findre2 = findre2.replace(".", "-")
        if findre2.find("日") == 4 or findre2.find("号") == 4:
            findre2 = findre2[:3] + "0" + findre2[3:]
        findre2 = findre2.replace("日", "")  # 12-02
        findre2 = findre2.replace("号", "")  # 12-02
        findre2 = datetime.date.today().__str__()[:5] + findre2  # 2018-12-02
        return findre2
    findre = re.findall(r'([123]?[0-9][日号])', dialog)  # 2号
    if len(findre) > 0:
        findre2 = findre[0]
        if findre2.find("日") == 1:
            findre2 = "0" + findre2  # 02号
        findre2 = findre2.replace("日", "")
        findre2 = findre2.replace("号", "")  # 02
        findre2 = datetime.date.today().__str__()[:8] + findre2  # 2018-03-02
        return findre2
    if -1 < dialog.find("今"):
        return datetime.date.today().__str__()
    elif -1 < dialog.find("明"):
        return (datetime.date.today() + datetime.timedelta(days=1)).__str__()
    elif -1 < dialog.find("后"):
        return (datetime.date.today() + datetime.timedelta(days=2)).__str__()
    return datetime.date.today()


def wash(strt):
    """
    先定位冒号
    从冒号往左找第一个不是数字的位置
    从冒号往右找第一个不是数字的位置
    截取之间的字符串返回
    :param strt:
    :return:
    """
    liststr = list(strt)
    safel = 0
    safer = len(liststr)
    mh = -1
    for i in range(safer):
        if liststr[i] == ":":
            mh = i
            break
    assert mh > safel
    for i in range(mh-1, safel-1, -1):
        if ord("0") <= ord(liststr[i]) <= ord("9"):
            continue
        safel = i + 1
        break
    for i in range(mh+1, safer, 1):
        if ord("0") <= ord(liststr[i]) <= ord("9"):
            continue
        safer = i + int(i == safer)
        break
    anstr = ''.join(liststr[safel:safer])
    return anstr


def create_sheet(sheetname, file):
    """
    根据sheetname创建sheet
    创建第一行
    :param sheetname:
    :return:
    """
    if sheetname in file.get_sheet_names():
        return
    file.create_sheet(sheetname)
    sheet = file.get_sheet_by_name(sheetname)
    sheet.cell(row=1, column=1).value = datetime.date.today().__str__()  # [:8]+"01"
    writetime(sheet=sheet, startrow=2)
    ccolumn = 1
    meeting_roomnames = ["12楼小会议室", "12楼大会议室", "13楼会议室"]
    for i in meeting_roomnames:
        ccolumn = ccolumn + 1
        sheet.cell(row=1, column=ccolumn).value = i
    return sheet


def get_excel_row(sheet, today):
    """
    :param sheet:
    :param today: datetime.date.today().__str__()
    :return: row of today
    """
    find_ornot = False
    find_row = 0
    for i in range(1, sheet.max_row+1):
        if sheet.cell(row=i, column=1).value == today:
            find_ornot = True
            find_row = i
            break
    if find_ornot:
        writetime(sheet=sheet, startrow=find_row+1)
        return find_row
    else:
        find_row = sheet.max_row + 1
        sheet.cell(row=find_row, column=1).value = datetime.date.today().__str__()
        writetime(sheet=sheet, startrow=find_row + 1)
        return find_row


def writetime(sheet, startrow):
    """
    不包括today信息的一天时间
    :param sheet:
    :param startrow:
    :return:
    """
    # excel_sheet.cell(row=startrow, excel_column=1).value = datetime.date.today().__str__()  # [:8]+"01"
    m = ["00", "15", "30", "45"]
    h = [i.__str__() for i in range(8, 21, 1)]
    crow = startrow
    for i in h:
        for j in m:
            sheet.cell(row=crow, column=1).value = i + ":" + j + ":00"
            crow = crow + 1


def occupy_it(sheet, st, en, co, info="占用人信息"):
    for i in range(st, en, 1):
        _ = sheet.cell(column=co, row=i, value=info)


def get_dtime(st, et):
    """
    计算所给时间段距离8:00的格子数 默认15min
    请注意8:00是第一个格子
    :param st:
    :param et:
    :return:
    """
    a = int(st[:st.find(":")])
    b = int(st[st.find(":") + 1:])
    c = int(et[:et.find(":")])
    d = int(et[et.find(":") + 1:])
    ds = (a - 8) * 4 + b // 15 + 1
    de = (c - 8) * 4 + d // 15 + 1
    return ds, de


def get_excel_column(mr):
    a = -1
    if mr.find("12楼小会议室") > -1:
        a = 2
    elif mr.find("12楼大会议室") > -1:
        a = 3
    elif mr.find("13楼会议室") > -1:
        a = 4
    return a


def get_excel_file(filename):
    """
    :param filename:
    :return:
    """
    dir_files = os.listdir(os.getcwd())
    if filename in dir_files:
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        wb.save(filename)
        wb = load_workbook(filename)
    return wb


def get_excel_sheet(riqi, file):
    sheetnames = file.get_sheet_names()  # 所有表名
    month_name = riqi[:7]
    if month_name in sheetnames:  # 存在
        sheet = file.get_sheet_by_name(name=month_name)
    else:
        sheet = create_sheet(month_name, file)
    return sheet


def is_occupied(sheet, start, end, column):
    busy = False  # 假设没占用
    busy_info = ""
    for i in range(start, end, 1):
        if sheet.cell(row=i, column=column).value is not None:
            busy_info = sheet.cell(row=i, column=column).value
            busy = True
            break
    return busy, busy_info


def deal_book(sheet, start, end, column, info, book, bot, contact):
    if book:
        occupied, occupied_info = is_occupied(sheet, start, end, column)
        if occupied:
            bot.SendTo(contact, "预定失败\"" + occupied_info + "\"")
            print("已被\"" + occupied_info + "\"占用")

        else:
            occupy_it(sheet, start, end, column, info)
            print("成功预定")
    else:  # todo 取消预定
        pass


def excel_file_close(file, name):
    file.save(filename=name)


watch_group_name = "0.0"
excel_file_name = "testMeeting.xlsx"


def onQQMessage(bot, contact, member, content):
    if not is_watch_group(contact=contact, group_name=watch_group_name):
        return
    if '[@ME]' in content:
        bot.SendTo(contact, "at命令未启用")
    dialog = dialog_clearify(content)
    dialog = is_cmd(dialog)
    if dialog is None:
        print("不是预定会议室")
    else:
        yuding_info = member.name
        book_ornot = find_yuding(dialog)  # True False
        riqi = find_riqi(dialog)  # 2018-03-13

        start_time, end_time = find_shijian(dialog)  # 8:00, 9:15
        if start_time is None or end_time is None:
            print("没找到时间")
        else:
            print("获取时间完成", start_time, end_time)
            fangjian = find_fangjian(dialog)  # 和昌12楼小会议室
            print("获取房间名完成")
            # 表格文件对象
            excel_file = get_excel_file(filename=excel_file_name)  # 检查是否被占用，，，
            print("获取文件", excel_file)
            excel_sheet = get_excel_sheet(riqi=riqi, file=excel_file)
            print("获取sheet", excel_sheet)
            excel_column = get_excel_column(fangjian)  # todo 没有就新建 找的时候在第一行寻找 异常处理
            print("获取列", excel_column)
            excel_date_row = get_excel_row(sheet=excel_sheet, today=riqi)  # 下一行是8:00:00
            print("获取行", excel_date_row)
            delta_row_start, delta_row_end = get_dtime(start_time, end_time)
            print("获取行区段", delta_row_start, delta_row_end)
            deal_book(sheet=excel_sheet, start=excel_date_row + delta_row_start,
                      end=excel_date_row + delta_row_end, column=excel_column,
                      info=yuding_info, book=book_ornot, bot=bot, contact=contact)

            # 存储表格文件
            excel_file.save(filename=excel_file_name)
            print("\n")



