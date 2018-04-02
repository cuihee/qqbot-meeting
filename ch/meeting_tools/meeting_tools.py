import datetime
from openpyxl import load_workbook, Workbook
import os
import re


def my_watch_group(contact, group_name):
    return contact.nick == group_name

# todo 考虑使用redis json存储规则
def dialog_clearify(content):
    """
    这里面的顺序不要乱
    :param content:
    :return:
    """
    d = content
    if not isinstance(d, type("字符串类型")):
        return ""
    clearify_dict = {'~~~': '-', "~~": "-", '  ': ' ',
                     '~~': '-', '~': '-', '---': '-',
                     '--': '-', '——': '-', '：': ':',
                     '全天': '8:30-18:00',
                     '12楼': '12层', '13楼': '13层',
                     '今日': '今天', '明日': '明天',
                     '合昌': '和昌', '合唱': '和昌', '和唱': '和昌',
                     '点半': ':30', '点30': ':30', '点三十': ':30',
                     '点': ':00',
                     '十一': '11', '十二': '12', '十三': '13',
                     '十四': '14', '十五': '15', '十六': '16',
                     '十七': '17', '十八': '18', '十九': '19',
                     '二十': '20', '十': '10',
                     '九': '9', '八': '8', '七': '7',
                     '六': '6', '五': '5', '四': '4',
                     '三': '3', '二': '2', '一': '1',
                     '下午1:': '13:', '下午2:': '14:', '下午3:': '15:', '下午4:': '16:',
                     '下午5:': '17:', '下午6:': '18:', '下午7:': '19:', '下午8:': '20:',
                     '预订': '预定',
                     '到': '-',
                     '（': '(', '）': ')',
                     '，': '.', ',': '.', '。': '.'
                     }
    for (k, v) in clearify_dict.items():
        d = d.replace(k, v)
    return d


def is_cmd(dialog):
    if not isinstance(dialog, type("")):
        return ""
    if dialog.find("预定") > -1 and dialog.find("会议室") > -1:
        return dialog


def get_meetingrooms_names():
    return ['小会议室', '大会议室', '13层会议室']


def find_fangjian(dialog):
    """
    用正则表达式从dialog中取出定义好的会议室名称
    http://www.runoob.com/regexp/regexp-metachar.html
    正则表达式教程↑

    :param dialog: 传入的字符串
    :return: 传出会议室名字
    """
    huiyishi = get_meetingrooms_names()

    meeting_room = '12层大会议室'
    r1 = re.findall(r'([1][23]层)', dialog)
    r2 = re.findall(r'([大小]?会议室)', dialog)
    if r1.__len__() > 0:
        meeting_room = r1[0] + meeting_room[3:]
    if r2.__len__() > 0:
        meeting_room = meeting_room[:3] + r2[0]
    for i in range(len(huiyishi)):
        if meeting_room.find(huiyishi[i]) > -1:
            return i  # 0 1 2
    return len(huiyishi)  # 3


def find_shijian(dialog):
    st = et = None
    r1 = re.findall(r'([0-9]?[0-9])[:点：]([0-5][0-9])', dialog)
    if r1.__len__() >= 1:  # 至少有一个时间
        st = r1[0][0] + ':' + r1[0][1]
    if r1.__len__() >= 2:  # 有两个时间
        et = r1[1][0] + ':' + r1[1][1]
    return st, et


def find_yuding(dialog):
    return not (-1 < dialog.find("取消") < 6)


def find_riqi(dialog):
    """

    :param dialog: 4.2
    :return: 2018-04-02
    """
    findre = re.findall(r'([12]?[0-9])[月.]([0-3]?[0-9])[日号]?', dialog)  # 12月02日 10.3
    if len(findre) == 1:
        if len(findre[0][0]) == 1:
            a = '0' + findre[0][0]
        else:
            a = findre[0][0]
        if len(findre[0][1]) == 1:
            b = '0' + findre[0][1]
        else:
            b = findre[0][1]
        return datetime.date.today().__str__()[:5] + a + '-' + b  # 2018-12-02
    findre = re.findall(r'([0-3]?[0-9])[日号]', dialog)  # 2号
    if len(findre) == 1:
        if len(findre[0]) == 1:
            a = '0' + findre[0]
        else:
            a = findre[0]
        return datetime.date.today().__str__()[:8] + a  # 2018-12-02
    if -1 < dialog.find("今"):
        return datetime.date.today().__str__()
    elif -1 < dialog.find("明"):
        return (datetime.date.today() + datetime.timedelta(days=1)).__str__()
    elif -1 < dialog.find("后"):
        return (datetime.date.today() + datetime.timedelta(days=2)).__str__()
    return datetime.date.today()



def get_excel_row(sheet, today):
    """
    :param sheet:
    :param today: datetime.date.today().__str__()
    :return: row of today 只有日期没时间的那一行
    """
    find_ornot = False
    find_row = 0
    for i in range(1, sheet.max_row+1):
        if sheet.cell(row=i, column=1).value == today:
            find_ornot = True
            find_row = i
            break
    if find_ornot:
        writetime(sheet=sheet, startrow=find_row + 1)
        return find_row
    else:
        find_row = sheet.max_row + 1
        sheet.cell(row=find_row, column=1).value = today
        writetime(sheet=sheet, startrow=find_row + 1)
        return find_row


def writetime(sheet, startrow):
    """
    不包括today信息的一天时间
    :param sheet:
    :param startrow:
    :return:
    """
    m = ["00", "15", "30", "45"]
    h = [i.__str__() for i in range(8, 21, 1)]
    crow = startrow
    for _h in h:
        for _m in m:
            sheet.cell(row=crow, column=1).value = _h + ":" + _m + ":00"
            crow = crow + 1


def get_dtime(st, et):
    """
    todo 感觉可以写成一个函数
    计算所给时间段距离8:00的格子数 默认15min
    请注意8:00是第一个格子
    :param st: 时间 8：00
    :param et: 时间 9：30
    :return: 起始时间所在行是当天日期所在行+ds
    """
    a = int(st[:st.find(":")])
    b = int(st[st.find(":") + 1:])
    c = int(et[:et.find(":")])
    d = int(et[et.find(":") + 1:])
    ds = (a - 8) * 4 + b // 15 + 1
    de = (c - 8) * 4 + d // 15 + 1
    return ds, de


def get_excel_file(filename):
    """
    :param filename: 文件名带后缀的
    :return: 打开指定文件名的文件对象
    """
    # 得到当前系统目录下的文件名列表
    dir_files = os.listdir(os.getcwd())
    # 当前路径下有filename文件
    if filename in dir_files:
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        wb.save(filename)
        wb = load_workbook(filename)
    return wb


def get_excel_sheet(riqi, file):
    sheetnames = file.get_sheet_names()  # 所有表名
    month_name = riqi[:7]  # 目标表名
    if month_name in sheetnames:  # 存在
        return file.get_sheet_by_name(name=month_name)
    else:
        return create_sheet(month_name, file)


def create_sheet(sheetname, file):
    """

    :param sheetname: string
    :param file: excel文件对象
    :return: sheet对象
    """
    # 如果excel文件中有这个名字的sheet 就直接返回这个sheet对象
    if sheetname in file.get_sheet_names():
        return file.get_sheet_by_name(sheetname)
    # 在excel文件中新建一个名为sheetname的sheet
    file.create_sheet(sheetname)
    sheet = file.get_sheet_by_name(sheetname)
    # 左上角 A1 写入今天的日期
    sheet.cell(row=1, column=1).value = datetime.date.today().__str__()  # [:8]+"01"
    # 写时间
    writetime(sheet=sheet, startrow=2)
    # 写会议室名字
    meeting_roomnames = get_meetingrooms_names()
    for i, n in meeting_roomnames:
        sheet.cell(row=1, column=i).value = n
    return sheet


def deal_book(sheet, start, end, column, info, book, bot, contact):
    if book:
        occupied, occupied_info = is_occupied(sheet, start, end, column)
        if occupied:
            bot.SendTo(contact, "预定失败\"" + occupied_info + "\"")
            print("已被\"" + occupied_info + "\"占用")
        else:
            occupy_it(sheet, start, end, column, info)
            print("成功预定")
    else:  # todo 取消预定
        pass


def is_occupied(sheet, start, end, column):
    busy = False  # 假设没占用
    busy_info = ""
    for i in range(start, end, 1):
        if sheet.cell(row=i, column=column).value is not None:
            busy_info = sheet.cell(row=i, column=column).value
            busy = True
            break
    return busy, busy_info


def occupy_it(sheet, st, en, co, info="占用人信息"):
    for i in range(st, en, 1):
        sheet.cell(column=co, row=i).value = info


def excel_file_close(file, name):
    file.save(filename=name)

