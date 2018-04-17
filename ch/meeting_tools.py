"""
这个文件应该放在全局可以import的位置供xl.py调用
比如说现在我的D:/anaconda/Lib
（xl.py按照要求放在了C:/Users/c2534/.qqbot-tmp/plugins）

有时候这个文件会不起作用，手动重启qqbot用命令 qq fresh-restart扫码

"""
import datetime
from openpyxl import load_workbook, Workbook
import os
import re


def ask_info(file, dates):
    info = ['无记录']
    # 打开这个文件
    f = get_excel_file(file)
    # 用dates[:7]找sheet
    fsheet = get_excel_sheet(dates, f)
    # 用dates找row
    frow = -1
    for i in range(1, fsheet.max_row+1):
        if fsheet.cell(row=i, column=1).value == dates:
            frow = i
            break
    if frow == -1:
        return info
    # 默认deltarow为
    deltafrow = 52  # 8:00-20:45=12*4+4=52
    for i in range(2, fsheet.max_column+1):  # 列
        for j in range(1+frow, deltafrow+1+frow):  # 行
            if fsheet.cell(row=j, column=i).value is None:
                continue
            if fsheet.cell(row=j, column=i).value not in info:
                info.append(str(fsheet.cell(row=j, column=i).value))

    # 关闭文件
    f.close()

    if len(info) > 1:
        info[:] = info[1:]  # 删除第一个元素
    return info


def my_watch_group(contact, group_name):
    return contact.nick in group_name


# todo 考虑使用redis json存储规则
# 但因为是py脚本运行 随时可以修改代码 需求不强烈
def dialog_clearify(content):
    """
    这里面的顺序不要乱
    :param content:
    :return:
    """
    d = content
    if not isinstance(d, type("字符串类型")):
        return ""
    clearify_dict = {'~~~': '-', "~~": "-", '  ': ' ',
                     '~~': '-', '~': '-', '---': '-',
                     '--': '-', '——': '-', '：': ':',
                     '～': '-',
                     '全天': '8:30-18:00',
                     '12楼': '12层', '13楼': '13层', '十三楼': '13层', '十二楼': '12层',
                     '今日': '今天', '明日': '明天',
                     '合昌': '和昌', '合唱': '和昌', '和唱': '和昌',
                     '点半': ':30', '点30': ':30', '点三十': ':30',
                     '点': ':00',
                     '十一': '11', '十二': '12', '十三': '13',
                     '十四': '14', '十五': '15', '十六': '16',
                     '十七': '17', '十八': '18', '十九': '19',
                     '二十': '20', '十': '10',
                     '九': '9', '八': '8', '七': '7',
                     '六': '6', '五': '5', '四': '4',
                     '三': '3', '二': '2', '一': '1',
                     '下午8:': '20:',
                     '预订': '预定',
                     '到': '-',
                     '（': '(', '）': ')',
                     '，': '.', ',': '.', '。': '.'
                     }
    for (k, v) in clearify_dict.items():
        d = d.replace(k, v)
    # 5:30-6:00 转换成 17:30-18:00
    r1 = re.findall(r'([^0-9:-]?)([1-7])(:[0-9]{2,3}-)([2-8])(:[0-9]{2,3})', d)
    if len(r1) >= 1:
        d = re.subn(r'([^0-9:-]?)([1-7])(:[0-9]{2,3}-)([2-8])(:[0-9]{2,3})',
                    r1[0][0] + str(int(r1[0][1]) + 12) + r1[0][2] + str(int(r1[0][3]) + 12) + r1[0][4],
                    d)[0]

    r1 = re.findall(r'([^0-9:])([0-9]{1,3})(-[0-9]{1,3}:[0-9]{2,3})', d)
    if len(r1) >= 1:
        d = re.subn(r'([^0-9:])([0-9]{1,3})(-[0-9]{1,3}:[0-9]{2,3})', r1[0][0] + r1[0][1] + ':00' + r1[0][2], d)[0]
    return d


def is_cmd(dialog):
    if not isinstance(dialog, type("")):
        return ''
    if "会议室" in dialog:
        if "预" in dialog or "订" in dialog or "定" in dialog:
            return dialog
    return ''


def get_meetingrooms_names():
    return ['小会议室', '大会议室', '13层会议室', '【会议室名称不详】无法正确记录']


def find_fangjian(dialog):
    """
    用正则表达式从dialog中取出定义好的会议室名称
    http://www.runoob.com/regexp/regexp-metachar.html
    正则表达式教程↑

    :param dialog: 传入的字符串
    :return: 传出会议室序号
    """
    huiyishi = get_meetingrooms_names()

    meeting_room = '12层大会议室'
    r1 = re.findall(r'([1][23]层)', dialog)
    r2 = re.findall(r'([大小]?会议室)', dialog)
    if r1.__len__() > 0:
        if r1[0] == '13层':  # 当前情况下13层只有一个会议室2
            return 2
        meeting_room = r1[0] + meeting_room[3:]
    if r2.__len__() > 0:
        meeting_room = meeting_room[:3] + r2[0]
    for i in range(len(huiyishi)):
        if meeting_room.find(huiyishi[i]) > -1:
            return i  # 0 1 2
    return len(huiyishi)-1  # 3


def find_shijian(dialog):
    st = '8:30'
    et = '18:00'
    r1 = re.findall(r'([0-9]?[0-9])[:点：]([0-5][0-9])', dialog)
    if r1.__len__() >= 1:  # 至少有一个时间
        st = r1[0][0] + ':' + shijian_fenzhong_round(r1[0][1])
    if r1.__len__() >= 2:  # 有两个时间
        et = shijian_round2(r1[1][0] + ':' + r1[1][1])
    if r1.__len__() <= 0:
        if dialog.find('上午') > -1:
            st, et = '8:30', '12:00'
        if dialog.find('下午') > -1:
            st, et = '14:00', '18:00'
    # 安全检查是否在早8:00-20:30之间
    if int(st[:st.find(':')]) < 8:
        st = '8' + st[st.find(':'):]
    if int(st[:st.find(':')]) > 20:
        st = '20' + st[st.find(':'):]
    if int(et[:et.find(':')]) < 8:
        et = '8' + et[et.find(':'):]
    if int(et[:et.find(':')]) > 20:
        et = '20' + et[et.find(':'):]

    return st, et


def shijian_fenzhong_round(s):
    if int(s) < 15:
        return '00'
    if int(s) < 30:
        return '15'
    if int(s) < 45:
        return '30'
    if int(s) < 60:
        return '45'
    return '45'


def shijian_round2(s):
    if int(s[-2:]) > 45:
        return str(int(s[:s.find(':')])+1) + ':00'
    if int(s[-2:]) > 30:
        return s[:s.find(':')] + ':45'
    if int(s[-2:]) > 15:
        return s[:s.find(':')] + ':30'
    if int(s[-2:]) > 0:
        return s[:s.find(':')] + ':15'
    if int(s[-2:]) == 0:
        return s
    return s


def find_yuding(dialog):
    return not ("取消" in dialog)


def find_riqi(dialog):
    """

    :param dialog: 4.2
    :return: 2018-04-02
    """
    findre = re.findall(r'([12]?[0-9])[月.]([0-3]?[0-9])[日号]?', dialog)  # 12月02日 10.3
    if len(findre) == 1:
        if len(findre[0][0]) == 1:
            a = '0' + findre[0][0]
        else:
            a = findre[0][0]
        if len(findre[0][1]) == 1:
            b = '0' + findre[0][1]
        else:
            b = findre[0][1]
        return datetime.date.today().__str__()[:5] + a + '-' + b  # 2018-12-02
    findre = re.findall(r'([0-3]?[0-9])[日号]', dialog)  # 2号
    if len(findre) == 1:
        if len(findre[0]) == 1:
            a = '0' + findre[0]
        else:
            a = findre[0]
        return datetime.date.today().__str__()[:8] + a  # 2018-12-02
    if -1 < dialog.find("今"):
        return datetime.date.today().__str__()
    elif -1 < dialog.find("明"):
        return (datetime.date.today() + datetime.timedelta(days=1)).__str__()
    elif -1 < dialog.find("后"):
        return (datetime.date.today() + datetime.timedelta(days=2)).__str__()
    return datetime.date.today().__str__()


def get_excel_row(sheet, today):
    """
    :param sheet:
    :param today: datetime.date.today().__str__()
    :return: row of today 只有日期没时间的那一行
    """
    find_ornot = False
    find_row = 0
    for i in range(1, sheet.max_row+1):
        if sheet.cell(row=i, column=1).value == today:
            find_ornot = True
            find_row = i
            break
    if find_ornot:
        writetime(sheet=sheet, startrow=find_row + 1)
        return find_row
    else:
        find_row = sheet.max_row + 1
        sheet.cell(row=find_row, column=1).value = today
        writetime(sheet=sheet, startrow=find_row + 1)
        return find_row


def writetime(sheet, startrow):
    """
    不包括today信息的一天时间
    :param sheet:
    :param startrow:
    :return:
    """
    m = ["00", "15", "30", "45"]
    h = [i.__str__() for i in range(8, 21, 1)]
    crow = startrow
    for _h in h:
        for _m in m:
            sheet.cell(row=crow, column=1).value = _h + ":" + _m + ":00"
            crow = crow + 1


def get_dtime(st, et):
    """
    todo 感觉可以写成一个函数
    计算所给时间段距离8:00的格子数 默认15min
    请注意8:00是第一个格子
    :param st: 时间 8：00
    :param et: 时间 9：30
    :return: 起始时间所在行是当天日期所在行+ds
    """
    a = int(st[:st.find(":")])
    b = int(st[st.find(":") + 1:])
    c = int(et[:et.find(":")])
    d = int(et[et.find(":") + 1:])
    ds = (a - 8) * 4 + b // 15 + 1
    de = (c - 8) * 4 + d // 15 + 1
    return ds, de


def get_excel_file(filename):
    """
    :param filename: 文件名带后缀的
    :return: 打开指定文件名的文件对象
    """
    # 得到当前系统目录下的文件名列表
    dir_files = os.listdir(os.getcwd())
    # 当前路径下有filename文件
    if filename in dir_files:
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        wb.save(filename)
        wb = load_workbook(filename)
    return wb


def get_excel_sheet(riqi, file):
    sheetnames = file.get_sheet_names()  # 所有表名
    month_name = riqi[:7]  # 目标表名
    if month_name in sheetnames:  # 存在
        return file.get_sheet_by_name(name=month_name)
    else:
        return create_sheet(month_name, file)


def create_sheet(sheetname, file):
    """

    :param sheetname: string
    :param file: excel文件对象
    :return: sheet对象
    """
    # 如果excel文件中有这个名字的sheet 就直接返回这个sheet对象
    if sheetname in file.get_sheet_names():
        return file.get_sheet_by_name(sheetname)
    # 在excel文件中新建一个名为sheetname的sheet
    file.create_sheet(sheetname)
    sheet = file.get_sheet_by_name(sheetname)
    # 左上角 A1 写入今天的日期
    sheet.cell(row=1, column=1).value = datetime.date.today().__str__()  # [:8]+"01"
    # 写时间
    writetime(sheet=sheet, startrow=2)
    # 写会议室名字
    meeting_roomnames = get_meetingrooms_names()
    # for (i, n) in meeting_roomnames:
    #     sheet.cell(row=1, column=i).value = n
    for i in range(len(meeting_roomnames)):
        sheet.cell(row=1, column=i+2).value = meeting_roomnames[i]
    return sheet


def deal_book(sheet, start, end, column, info, book, bot, contact, member):
    if book:
        # 预定命令
        occupied, occupied_info = is_occupied(sheet, start, end, column)  # 是否被占用 占用信息
        if occupied:
            # 如果占用
            bot.SendTo(contact, "机器人回复 失败，因为\"" + occupied_info + "\"占用")
            # print("您预定失败，因为\"" + occupied_info + "\"占用")
        else:
            # 没有占用
            occupy_it(sheet, start, end, column, info)
            bot.SendTo(contact, "机器人回复 成功\n"+" 记录的信息： "+member.name+" "+info[-32:])
            # print("成功预定")  # todo 谁 预定成功了 日期 时间 房间
    else:
        # 取消预定
        unoccupy_it(sheet, start, end, column)
        bot.SendTo(contact, '机器人回复 '+str(info[:info.find(' 群"')]) + "取消成功")
        # print("取消预定")
    print('\n')


def is_occupied(sheet, start, end, column):
    busy = False  # 假设没占用
    busy_info = ""
    for i in range(start, end, 1):
        if sheet.cell(row=i, column=column).value is not None:
            busy_info = sheet.cell(row=i, column=column).value
            busy = True
            break
    return busy, busy_info


def occupy_it(sheet, st, en, co, info="占用人信息"):
    for i in range(st, en, 1):
        sheet.cell(column=co, row=i).value = info


def unoccupy_it(sheet, st, en, co):
    for i in range(st, en, 1):
        sheet.cell(column=co, row=i).value = None


def excel_file_close(file, name):
    file.save(filename=name)


def _test_dialog_clearify():
    assert dialog_clearify('预定4月2日和昌12楼小会议室5:30-6:00') \
        == '预定4月2日和昌12层小会议室17:30-18:00', '对话语句清理函数dialog_clearify有问题'
    assert dialog_clearify('预定4月2日和昌12楼小会议室2:30-6:00') \
        == '预定4月2日和昌12层小会议室14:30-18:00', '对话语句清理函数dialog_clearify有问题'
    assert dialog_clearify('预定4月3日和昌12楼大会议室9:00-11：00') \
        == '预定4月3日和昌12层大会议室9:00-11:00', '对话语句清理函数dialog_clearify有问题'
    assert dialog_clearify('预定4月2日上午和昌12楼小会议室，9:00--11:30') \
        == '预定4月2日上午和昌12层小会议室.9:00-11:30', '对话语句清理函数dialog_clearify有问题'
    assert dialog_clearify('预定28日（今天）下午12楼大会议室，15:00到16:00') \
        == '预定28日(今天)下午12层大会议室.15:00-16:00', '对话语句清理函数dialog_clearify有问题'
    assert dialog_clearify('预定今天全天小会议室') \
        == '预定今天8:30-18:00小会议室', '对话语句清理函数dialog_clearify有问题'
    assert dialog_clearify('预定4月2日和昌13楼大会议室11:00-12:00') \
        == '预定4月2日和昌13层大会议室11:00-12:00', '对话语句清理函数dialog_clearify有问题'
    assert dialog_clearify('预定27号，下午14点30到17点，和昌12层大会议室') \
        == '预定27号.下午14:30-17:00.和昌12层大会议室', '对话语句清理函数dialog_clearify有问题'


def _test_find_fangjian():
    assert find_fangjian('预定4月2日和昌12层小会议室17:30-18:00') == 0, '寻找会议室编号函数find_fangjian有问题'
    assert find_fangjian('预定4月3日和昌12层大会议室9:00-11:00') == 1, '寻找会议室编号函数find_fangjian有问题'
    assert find_fangjian('预定4月2日和昌13层大会议室11:00-12:00') == 2, '寻找会议室编号函数find_fangjian有问题'
    assert find_fangjian('预定4月2日和昌13层会议室11:00-12:00') == 2, '寻找会议室编号函数find_fangjian有问题'
    assert find_fangjian('预定今天8:30-18:00小会议室') == 0, '寻找会议室编号函数find_fangjian有问题'
    assert find_fangjian('预定今天8:30-18:00老楼会议室') == 3, '寻找会议室编号函数find_fangjian有问题'


def _test_find_shijian():
    assert find_shijian('预定4月2日和昌12层小会议室17:30-18:00') == ('17:30', '18:00'), '寻找开始和结束时间函数find_shijian有问题'
    assert find_shijian('预定4月2日和昌12层小会议室14:30-18:00') == ('14:30', '18:00'), '寻找开始和结束时间函数find_shijian有问题'
    assert find_shijian('预定28日(今天)下午12层大会议室.15:00-16:00') == ('15:00', '16:00'), '寻找开始和结束时间函数find_shijian有问题'
    assert find_shijian('预定今天8:30-18:00小会议室') == ('8:30', '18:00'), '寻找开始和结束时间函数find_shijian有问题'
    assert find_shijian('预定上午大会议室') == ('8:30', '12:00'), '寻找开始和结束时间函数find_shijian有问题'
    assert find_shijian('订小会议室 9:10-10:30') == ('9:00', '10:30'), '寻找开始和结束时间函数find_shijian有问题'


def _test_find_riqi():
    assert find_riqi('预定4月2日和昌12层小会议室17:30-18:00') == '2018-04-02', '寻找日期函数find_riqi有问题'
    assert find_riqi('预定今天8:30-18:00小会议室') == datetime.date.today().__str__(), '寻找日期函数find_riqi有问题'
    assert find_riqi('预定30日下午12楼小会议室.14:00-16:00') == '2018-04-30', '寻找日期函数find_riqi有问题'
    assert find_riqi('预定27号.下午14:30-17:00.和昌12层大会议室') == '2018-04-27', '寻找日期函数find_riqi有问题'
    assert find_riqi('预订12楼大会议室，12点到14点')[:7] == datetime.date.today().__str__()[:7], '寻找日期函数find_riqi有问题'


if __name__ == '__main__':
    _test_dialog_clearify()
    _test_find_fangjian()
    _test_find_shijian()
    _test_find_riqi()
